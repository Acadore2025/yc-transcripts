"""
YC YouTube Transcript Downloader — Excel Mode (YouTube API v3)
==============================================================
Uses YouTube Data API v3 for metadata (no IP blocking),
and youtube-transcript-api for transcripts.
Reads URLs from Excel, saves PDFs, updates status back to Excel.

SETUP:
    pip install requests youtube-transcript-api reportlab openpyxl

API KEY SETUP (one time):
    Windows VS Code terminal:
        setx YOUTUBE_API_KEY "AIzaSyxxxxxxxxxxxxxxxxxxxxxxxxx"
        → Restart VS Code after running this

    OR create a .env file in the same folder as this script:
        YOUTUBE_API_KEY=AIzaSyxxxxxxxxxxxxxxxxxxxxxxxxx

RUN:
    python yc_transcript_downloader.py
    → Enter path to your Excel file when prompted

EXCEL FORMAT (yc_urls_template.xlsx):
    A: YouTube URL
    B: Folder Name
    C: Status       ← script writes this
    D: Error Reason ← script writes this
    E: Downloaded At← script writes this
"""

import re
import os
import sys
import time
import logging
import requests
from pathlib import Path
from datetime import datetime

from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from youtube_transcript_api import YouTubeTranscriptApi
from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, HRFlowable

# ─── CONFIG ───────────────────────────────────────────────────────────────────

SLEEP_SECS     = 1.0   # seconds between requests — polite to YouTube API
STATUS_COL     = 3     # C
ERROR_COL      = 4     # D
TIMESTAMP_COL  = 5     # E

# YouTube Data API v3 base URL
YT_API_BASE = "https://www.googleapis.com/youtube/v3"

# Status values
S_PENDING   = "Pending"
S_DONE      = "Done"
S_FAILED    = "Failed"
S_DUPLICATE = "Skipped (duplicate)"

# Cell colours
FILL_DONE  = PatternFill("solid", start_color="E2EFDA")
FILL_FAIL  = PatternFill("solid", start_color="FCE4D6")
FILL_SKIP  = PatternFill("solid", start_color="FFF2CC")
FILL_CLEAR = PatternFill(fill_type=None)

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s  %(levelname)-8s  %(message)s",
)
log = logging.getLogger(__name__)


# ─── API KEY LOADER ───────────────────────────────────────────────────────────

def load_api_key() -> str:
    """
    Load YouTube API key from:
    1. Environment variable YOUTUBE_API_KEY
    2. .env file in same folder as script
    3. Manual input as fallback
    """
    # 1. Check environment variable (set via setx on Windows)
    key = os.environ.get("YOUTUBE_API_KEY", "").strip()
    if key:
        log.info("✅ API key loaded from environment variable")
        return key

    # 2. Check .env file
    env_file = Path(__file__).parent / ".env"
    if env_file.exists():
        for line in env_file.read_text().splitlines():
            if line.startswith("YOUTUBE_API_KEY="):
                key = line.split("=", 1)[1].strip().strip('"').strip("'")
                if key:
                    log.info("✅ API key loaded from .env file")
                    return key

    # 3. Ask user to paste it
    print("\n⚠️  YouTube API key not found in environment.")
    print("   You can set it permanently by running in terminal:")
    print('   setx YOUTUBE_API_KEY "your-key-here"  (then restart VS Code)')
    print("\n   OR paste it now for this session only:")
    key = input("   YouTube API Key: ").strip()
    return key


# ─── YOUTUBE API HELPERS ──────────────────────────────────────────────────────

def extract_video_id(url: str) -> str:
    """Extract YouTube video ID from any YouTube URL format."""
    match = re.search(r"(?:v=|youtu\.be/|embed/|shorts/)([a-zA-Z0-9_-]{11})", url)
    if not match:
        raise ValueError("Invalid YouTube URL — could not extract video ID")
    return match.group(1)


def get_video_metadata(video_id: str, api_key: str) -> dict:
    """
    Fetch video metadata using YouTube Data API v3.
    This uses Google's servers — no IP blocking.
    """
    try:
        url = f"{YT_API_BASE}/videos"
        params = {
            "part": "snippet,contentDetails",
            "id": video_id,
            "key": api_key,
        }
        resp = requests.get(url, params=params, timeout=10)
        resp.raise_for_status()
        data = resp.json()

        if not data.get("items"):
            return _empty_metadata(video_id)

        item = data["items"][0]
        snippet = item.get("snippet", {})
        duration_raw = item.get("contentDetails", {}).get("duration", "")

        # Parse ISO 8601 duration (PT1H2M3S → seconds)
        duration_secs = parse_iso_duration(duration_raw)

        # Parse publish date
        published = snippet.get("publishedAt", "")[:10]  # YYYY-MM-DD

        return {
            "title":          snippet.get("title", "Unknown Title"),
            "channel":        snippet.get("channelTitle", "Y Combinator"),
            "upload_date":    published,
            "duration":       duration_secs,
            "url":            f"https://www.youtube.com/watch?v={video_id}",
            "playlist_title": "",
            "description":    snippet.get("description", "")[:300],
        }

    except requests.exceptions.HTTPError as e:
        if e.response.status_code == 403:
            log.error("❌ API key invalid or quota exceeded. Check your key.")
        raise
    except Exception as e:
        log.warning(f"Metadata fetch failed: {e}")
        return _empty_metadata(video_id)


def _empty_metadata(video_id: str) -> dict:
    return {
        "title": f"Video_{video_id}",
        "channel": "Y Combinator",
        "upload_date": "",
        "duration": 0,
        "url": f"https://www.youtube.com/watch?v={video_id}",
        "playlist_title": "",
        "description": "",
    }


def parse_iso_duration(duration: str) -> int:
    """Convert ISO 8601 duration (PT1H2M3S) to seconds."""
    if not duration:
        return 0
    pattern = r"PT(?:(\d+)H)?(?:(\d+)M)?(?:(\d+)S)?"
    match = re.match(pattern, duration)
    if not match:
        return 0
    hours   = int(match.group(1) or 0)
    minutes = int(match.group(2) or 0)
    seconds = int(match.group(3) or 0)
    return hours * 3600 + minutes * 60 + seconds


# ─── TRANSCRIPT FETCHER ───────────────────────────────────────────────────────

def fetch_transcript(video_id: str) -> tuple[str, str]:
    """
    Fetch transcript via youtube-transcript-api.
    Returns (text, source_label).
    """
    # Try English first
    try:
        api = YouTubeTranscriptApi()
        fetched = api.fetch(video_id, languages=["en", "en-US", "en-GB"])
        return format_transcript(list(fetched)), "manual/auto"
    except Exception:
        pass

    # Fallback: any available language
    try:
        api = YouTubeTranscriptApi()
        fetched = api.fetch(video_id)
        return format_transcript(list(fetched)), "auto"
    except Exception as e:
        return "", f"No transcript: {e}"


def format_transcript(entries) -> str:
    """Group entries into clean ~60 second paragraphs."""
    lines, buffer, buf_dur = [], [], 0.0
    for entry in entries:
        text     = entry.text     if hasattr(entry, "text")     else entry.get("text", "")
        duration = entry.duration if hasattr(entry, "duration") else entry.get("duration", 0)
        text = re.sub(r"\[.*?\]", "", text).replace("\n", " ").strip()
        if not text:
            continue
        buffer.append(text)
        buf_dur += duration
        if buf_dur >= 60:
            lines.append(" ".join(buffer))
            buffer, buf_dur = [], 0.0
    if buffer:
        lines.append(" ".join(buffer))
    return "\n\n".join(lines)


# ─── EXCEL HELPERS ────────────────────────────────────────────────────────────

def load_rows(xlsx_path: Path):
    wb = load_workbook(xlsx_path)
    ws = wb["Transcripts"]
    rows = []
    for row in ws.iter_rows(min_row=2):
        url    = (row[0].value or "").strip()
        folder = (row[1].value or "General").strip()
        status = (row[2].value or S_PENDING).strip()
        if not url:
            continue
        rows.append({
            "row_num": row[0].row,
            "url":     url,
            "folder":  folder,
            "status":  status,
        })
    return wb, ws, rows


def update_row(ws, row_num: int, status: str, error: str = "", timestamp: str = ""):
    ws.cell(row=row_num, column=STATUS_COL).value    = status
    ws.cell(row=row_num, column=ERROR_COL).value     = error
    ws.cell(row=row_num, column=TIMESTAMP_COL).value = timestamp
    fill = {S_DONE: FILL_DONE, S_FAILED: FILL_FAIL,
            S_DUPLICATE: FILL_SKIP}.get(status, FILL_CLEAR)
    for col in range(1, 6):
        ws.cell(row=row_num, column=col).fill = fill


# ─── PDF BUILDER ──────────────────────────────────────────────────────────────

def sanitize_filename(name: str, max_len: int = 80) -> str:
    name = re.sub(r'[\\/*?:"<>|#%&{}\$!\'@+`=]', "", name)
    name = re.sub(r"\s+", "_", name.strip())
    return name[:max_len]


def save_as_pdf(output_path: Path, video: dict, transcript: str, source: str, folder: str):
    doc = SimpleDocTemplate(str(output_path), pagesize=letter,
                            rightMargin=inch, leftMargin=inch,
                            topMargin=inch, bottomMargin=inch)
    styles = getSampleStyleSheet()

    title_style = ParagraphStyle("T", parent=styles["Title"], fontSize=18,
                                 spaceAfter=6, textColor=colors.HexColor("#FF6600"))
    meta_style  = ParagraphStyle("M", parent=styles["Normal"], fontSize=9,
                                 textColor=colors.HexColor("#888888"), spaceAfter=4)
    body_style  = ParagraphStyle("B", parent=styles["Normal"], fontSize=11,
                                 leading=16, spaceAfter=10)

    def safe(t):
        return (t or "").replace("&","&amp;").replace("<","&lt;").replace(">","&gt;")

    story = [Paragraph(safe(video["title"]), title_style)]

    meta_items = [
        f"<b>URL:</b> {safe(video['url'])}",
        f"<b>Channel:</b> {safe(video.get('channel', 'Y Combinator'))}",
        f"<b>Folder:</b> {safe(folder)}",
    ]
    if video.get("upload_date"):
        meta_items.append(f"<b>Published:</b> {safe(video['upload_date'])}")
    if video.get("duration"):
        m, s = divmod(int(video["duration"]), 60)
        meta_items.append(f"<b>Duration:</b> {m}m {s}s")
    meta_items.append(f"<b>Transcript source:</b> {source}")
    meta_items.append(f"<b>Generated:</b> {datetime.now().strftime('%Y-%m-%d %H:%M')}")

    for item in meta_items:
        story.append(Paragraph(item, meta_style))

    story += [Spacer(1, 8),
              HRFlowable(width="100%", thickness=1, color=colors.HexColor("#FF6600")),
              Spacer(1, 12)]

    if transcript:
        for para in transcript.split("\n\n"):
            para = para.strip()
            if para:
                story.append(Paragraph(safe(para), body_style))
    else:
        story.append(Paragraph("<i>No transcript available for this video.</i>", body_style))

    doc.build(story)


# ─── MAIN ─────────────────────────────────────────────────────────────────────

def main():
    print("=" * 60)
    print("  YC Transcript Downloader — YouTube API Mode")
    print("=" * 60)

    # Load API key
    api_key = load_api_key()
    if not api_key:
        print("❌ No API key provided. Exiting.")
        return

    # Get Excel path
    if len(sys.argv) > 1:
        xlsx_path = Path(sys.argv[1])
    else:
        raw = input("\nPath to Excel file (e.g. yc_urls_template.xlsx): ").strip().strip('"')
        xlsx_path = Path(raw)

    if not xlsx_path.exists():
        print(f"❌ File not found: {xlsx_path}")
        return

    # Output dir lives next to Excel file
    OUTPUT_DIR = xlsx_path.parent / "yc_transcripts"
    OUTPUT_DIR.mkdir(exist_ok=True)
    log.info(f"Output directory: {OUTPUT_DIR}")

    # Load Excel rows
    wb, ws, rows = load_rows(xlsx_path)
    log.info(f"Loaded {len(rows)} rows from {xlsx_path.name}")

    # Detect duplicates
    seen_urls = {}
    for r in rows:
        url = r["url"].lower()
        r["is_duplicate"] = url in seen_urls
        seen_urls[url] = True

    counts = {S_DONE: 0, S_FAILED: 0, S_DUPLICATE: 0, "skipped_done": 0}

    for idx, row in enumerate(rows, 1):
        url     = row["url"]
        folder  = row["folder"]
        status  = row["status"]
        row_num = row["row_num"]

        print(f"\n[{idx}/{len(rows)}] {url[:70]}")

        # Skip duplicates
        if row["is_duplicate"]:
            log.info("  → Duplicate, skipping")
            update_row(ws, row_num, S_DUPLICATE, "Same URL already in sheet")
            wb.save(xlsx_path)
            counts[S_DUPLICATE] += 1
            continue

        # Skip already done
        if status == S_DONE:
            log.info("  → Already done, skipping")
            counts["skipped_done"] += 1
            continue

        # 1. Extract video ID
        try:
            video_id = extract_video_id(url)
        except ValueError as e:
            log.warning(f"  → {e}")
            update_row(ws, row_num, S_FAILED, str(e), datetime.now().strftime("%Y-%m-%d %H:%M"))
            wb.save(xlsx_path)
            counts[S_FAILED] += 1
            continue

        # 2. Metadata via YouTube API (no IP blocking)
        log.info("  → Fetching metadata via YouTube API...")
        try:
            video = get_video_metadata(video_id, api_key)
            log.info(f"  → Title: {video['title']}")
        except Exception as e:
            log.warning(f"  → Metadata failed: {e}")
            video = _empty_metadata(video_id)
            video["url"] = url

        # 3. Transcript
        log.info("  → Fetching transcript...")
        transcript, source = fetch_transcript(video_id)

        if not transcript:
            log.warning(f"  → No transcript: {source}")
            update_row(ws, row_num, S_FAILED, source, datetime.now().strftime("%Y-%m-%d %H:%M"))
            wb.save(xlsx_path)
            counts[S_FAILED] += 1
            continue

        log.info(f"  → Transcript OK ({len(transcript)} chars, {source})")

        # 4. Save PDF
        folder_path = OUTPUT_DIR / sanitize_filename(folder)
        folder_path.mkdir(parents=True, exist_ok=True)
        pdf_path = folder_path / (sanitize_filename(video["title"]) + ".pdf")

        try:
            save_as_pdf(pdf_path, video, transcript, source, folder)
            log.info(f"  → Saved: {pdf_path.name}")
            update_row(ws, row_num, S_DONE, "", datetime.now().strftime("%Y-%m-%d %H:%M"))
            counts[S_DONE] += 1
        except Exception as e:
            log.error(f"  → PDF error: {e}")
            update_row(ws, row_num, S_FAILED, f"PDF error: {e}", datetime.now().strftime("%Y-%m-%d %H:%M"))
            counts[S_FAILED] += 1

        wb.save(xlsx_path)
        time.sleep(SLEEP_SECS)

    # Summary
    print("\n" + "=" * 60)
    print("  COMPLETE")
    print(f"  ✅  Done          : {counts[S_DONE]}")
    print(f"  ❌  Failed        : {counts[S_FAILED]}")
    print(f"  ⏭️   Duplicates    : {counts[S_DUPLICATE]}")
    print(f"  ⏩  Already done  : {counts['skipped_done']}")
    print(f"  📁  Output folder : {OUTPUT_DIR.resolve()}")
    print("=" * 60)


if __name__ == "__main__":
    main()
